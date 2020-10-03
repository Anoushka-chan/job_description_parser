import os
import requests
import json
import xlrd
import openpyxl
from dotenv import load_dotenv
load_dotenv()

# wb = openpyxl.load_workbook(os.getenv('BASE_ROOT_PATH') + '/tags_with_alias.xlsx')
# sheet = wb.active
# c1 = sheet.cell(row = 3, column = 1) 
# c1.value = "Information Technology"
# c2 = sheet.cell(row= 3 , column = 2)
# c2.value = "I.T."
# wb.save(os.getenv('BASE_ROOT_PATH') + '/tags_with_alias.xlsx')




# file_path = os.getenv('BASE_ROOT_PATH') + '/tags'
# files = [os.path.join(file_path, fi) for fi in os.listdir('tags')]

loc = os.getenv('BASE_ROOT_PATH') + '/tags_with_alias.xlsx'
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
end_row = sheet.nrows

for iterator in range(1,end_row):
	word = sheet.cell_value(iterator, 0)
	alias = sheet.cell_value(iterator, 1)
	word = word.rstrip()
	word = word.lstrip()
	payload = {
		"index":"jd-hardskills",
		"type":"hardskills",
		"name":word ,
		"name-alias":alias
	}

	url = os.getenv('JD_HOST') + ':' + os.getenv('JD_PORT') + '/postsearch'
	headers = {'content-type': 'application/json'}
	response = requests.post(url, data=json.dumps(payload), headers=headers)
	results = json.loads(response.text)
	print('==============', results)


# for fil_num, fil in enumerate(files):
# 	with open(fil, errors='ignore') as fi:
# 		clean = fi.read()
# 		temp = clean.split(',')
# 		for word in temp:
# 			word = word.rstrip()
# 			word = word.lstrip()
# 			# print(word)   HIT API
# 			payload = {
# 				"index":"jd-hardskills",
# 				"type":"hardskills",
# 				"name":word ,
# 				"name-alias":input_alias
# 			}

# 			url = 'http://0.0.0.0:4997/postsearch'

# 			headers = {'content-type': 'application/json'}
# 			response = requests.post(url, data=json.dumps(payload), headers=headers)
# 			results = json.loads(response.text)
# 			print(results)
# 			#return jsonify({'esput': results})
			